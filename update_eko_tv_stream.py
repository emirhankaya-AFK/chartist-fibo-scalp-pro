import openpyxl
import json

eko_tv_stocks = [
    {
        "ticker": "SKBNK",
        "name": "Şekerbank T.A.Ş.",
        "entry": 7.50,
        "support": "5.00 - 6.00 TL",
        "resistance": "14.00 TL (Haftalık HO)",
        "note": "[Ahmet Mergen Eko TV] 20 TL'den 7.50 TL'ye düştü. Taban çözülme süreci takip edilmeli, 5.00-6.00 TL dip oluşum alanı."
    },
    {
        "ticker": "EUREN",
        "name": "Europen Endüstri İnşaat Sanayi A.Ş.",
        "entry": 3.80,
        "support": "3.50 TL (Dolar: 0.08 USD Halka Arz Dibi)",
        "resistance": "4.50 - 5.00 TL (20 & 50 Günlük HO)",
        "note": "[Ahmet Mergen Eko TV] Dolar bazında halka arz seviyesi olan 0.08 USD (8 cent) dibine geldi. 20 ve 50 günlük HO üzeri hareket takip edilmeli."
    },
    {
        "ticker": "TMSN",
        "name": "Tümosan Motor ve Traktör Sanayi A.Ş.",
        "entry": 86.00,
        "support": "80.00 TL (Ana Taban)",
        "resistance": "89.00 TL (50G HO), 105.00 - 110.00 TL (Kanal)",
        "note": "[Ahmet Mergen Eko TV] 80 TL'den tepki verdi. Sağ omuz TOBO oluşumu var. 50 günlük HO (89 TL) aşılırsa kanal hedefi 105-110 TL."
    },
    {
        "ticker": "KCHOL",
        "name": "Koç Holding A.Ş.",
        "entry": 190.00,
        "support": "185.00 TL (50 Haftalık HO) / 175.00 - 180.00 TL (%50 Fibo)",
        "resistance": "211.00 TL - 250.00 TL (Çanak Hedefi: 300+ TL)",
        "note": "[Ahmet Mergen Eko TV] 20 haftalık HO (193 TL) delindi. 50 haftalık HO (185 TL) ve 175-180 TL %50 Fibo ana destek bölgesi. Uzun vade hedef 300+ TL."
    },
    {
        "ticker": "ASELS",
        "name": "Aselsan Elektronik Sanayi A.Ş.",
        "entry": 361.75,
        "support": "300.00 - 320.00 TL (Ana Destek)",
        "resistance": "388.00 TL - 406.00 TL",
        "note": "[Ahmet Mergen Eko TV] Dikkatli olunmalı, stopsuz alım yapılmamalı. Geri çekilmelerde 300-320 TL ana alım/destek bölgesi."
    }
]

# 1. Update Hisselerin_Teknik_Verileri.xlsx
wb = openpyxl.load_workbook('Hisselerin_Teknik_Verileri.xlsx')
ws = wb.active

existing_rows = {}
for r in range(4, ws.max_row + 1):
    val = ws.cell(row=r, column=2).value
    if val:
        ticker_str = str(val).upper().strip()
        existing_rows[ticker_str] = r

for stock in eko_tv_stocks:
    ticker = stock["ticker"]
    if ticker in existing_rows:
        row_idx = existing_rows[ticker]
        ws.cell(row=row_idx, column=4).value = stock["entry"]
        ws.cell(row=row_idx, column=5).value = stock["support"]
        ws.cell(row=row_idx, column=6).value = stock["resistance"]
        ws.cell(row=row_idx, column=8).value = stock["note"]
    else:
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = ticker
        ws.cell(row=new_row, column=2).value = ticker
        ws.cell(row=new_row, column=3).value = stock["name"]
        ws.cell(row=new_row, column=4).value = stock["entry"]
        ws.cell(row=new_row, column=5).value = stock["support"]
        ws.cell(row=new_row, column=6).value = stock["resistance"]
        ws.cell(row=new_row, column=8).value = stock["note"]

wb.save('Hisselerin_Teknik_Verileri.xlsx')
print("Hisselerin_Teknik_Verileri.xlsx updated with Eko TV Mergen stream.")

# 2. Update manual_tracking.json
with open("manual_tracking.json", "r", encoding="utf-8") as f:
    manual_list = json.load(f)

for stock in eko_tv_stocks:
    if stock["ticker"] not in manual_list:
        manual_list.append(stock["ticker"])

with open("manual_tracking.json", "w", encoding="utf-8") as f:
    json.dump(manual_list, f, ensure_ascii=False, indent=2)
print("manual_tracking.json updated.")
