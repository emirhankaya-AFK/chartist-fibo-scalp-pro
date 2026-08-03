"""15-minute delayed opportunity poller.

It never places orders. Notifications are opt-in via NTFY_TOPIC and NTFY_TOKEN.
"""
from __future__ import annotations

import json
import os
import time
import urllib.request
from pathlib import Path
from datetime import datetime, timezone, timedelta
from threading import Lock

from market_scanner import scan_market

TR_TZ = timezone(timedelta(hours=3))

def get_tr_now() -> datetime:
    """Returns current Turkey local time (UTC+3)."""
    return datetime.now(TR_TZ)


INTERVAL_SECONDS = int(os.getenv("OPPORTUNITY_INTERVAL_SECONDS", "900"))
STATE_PATH = Path(__file__).with_name("intraday_alert_state.json")
NOTIFICATION_LOG_PATH = Path(__file__).with_name("notification_log.json")
TRACKING_LOG_PATH = Path(__file__).with_name("tracking_log.json")
MANUAL_TRACKING_PATH = Path(__file__).with_name("manual_tracking.json")
OGUZ_ARSIV_PATH = Path(__file__).with_name("OGUZ_ANALIZ_ARSIVI.json")
EXCEL_PATH = Path(__file__).with_name("Hisselerin_Teknik_Verileri.xlsx")
DEFAULT_NTFY_TOPIC = "emirkan_bist_alarm"
_status_lock = Lock()
_status = {
    "running": False,
    "lastRun": None,
    "lastError": None,
    "opportunities": [],
    "tracking": [],
    "notificationsEnabled": True,
    "topic": os.getenv("NTFY_TOPIC", DEFAULT_NTFY_TOPIC).strip(),
}

def _load_state() -> dict[str, str]:
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}

def _save_state(state: dict[str, str]) -> None:
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=False), encoding="utf-8")


def manual_tracking() -> list[str]:
    try:
        rows = json.loads(MANUAL_TRACKING_PATH.read_text(encoding="utf-8")) if MANUAL_TRACKING_PATH.exists() else []
        return [str(item).upper().replace(".IS", "") for item in rows if str(item).strip()] if isinstance(rows, list) else []
    except (OSError, ValueError, TypeError):
        return []


def add_manual_tracking(ticker: str) -> list[str]:
    ticker = str(ticker).upper().replace(".IS", "").strip()
    rows = manual_tracking()
    if ticker and ticker not in rows:
        rows.append(ticker)
        MANUAL_TRACKING_PATH.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")
    return rows


def _append_notification_log(item: dict, message: str, delivered: bool) -> None:
    try:
        rows = json.loads(NOTIFICATION_LOG_PATH.read_text(encoding="utf-8")) if NOTIFICATION_LOG_PATH.exists() else []
        if not isinstance(rows, list):
            rows = []
        if not delivered:
            for previous in rows:
                if previous.get("ticker") == item.get("ticker") and previous.get("status") == "not_sent":
                    previous["timestamp"] = datetime.now().astimezone().isoformat(timespec="seconds")
                    previous["message"] = message
                    previous["attempts"] = int(previous.get("attempts", 1)) + 1
                    NOTIFICATION_LOG_PATH.write_text(json.dumps(rows[:500], ensure_ascii=False), encoding="utf-8")
                    return
        rows.insert(0, {
            "timestamp": datetime.now().astimezone().isoformat(timespec="seconds"),
            "ticker": item.get("ticker"),
            "price": item.get("price"),
            "strategy": item.get("strategy"),
            "score": item.get("score"),
            "message": message,
            "status": "sent" if delivered else "not_sent",
            "reason": "ntfy gönderimi başarısız veya NTFY_TOPIC eksik" if not delivered else "",
            "attempts": 1,
        })
        NOTIFICATION_LOG_PATH.write_text(json.dumps(rows[:500], ensure_ascii=False), encoding="utf-8")
    except (OSError, ValueError, TypeError):
        pass


def is_notification_window_open() -> bool:
    """Returns True ONLY if Turkey Local Time (UTC+3) is between 09:50 and 18:15 on weekdays."""
    now = get_tr_now()
    if now.weekday() >= 5:  # Weekend
        return False
    start_time = now.replace(hour=9, minute=50, second=0, microsecond=0)
    end_time = now.replace(hour=18, minute=15, second=0, microsecond=0)
    return start_time <= now <= end_time


def _notify(message: str) -> bool:
    if not is_notification_window_open():
        return False
    sent = False
    topic = os.getenv("NTFY_TOPIC", DEFAULT_NTFY_TOPIC).strip()
    if topic:
        url = f"https://ntfy.sh/{topic}"
        headers = {"Title": "Chartist Fibo-Scalp Pro", "Priority": "high", "Tags": "chart_with_upwards_trend,bell"}
        token = os.getenv("NTFY_TOKEN", "").strip()
        if token:
            headers["Authorization"] = f"Bearer {token}"
        request = urllib.request.Request(url, data=message.encode("utf-8"), headers=headers, method="POST")
        try:
            with urllib.request.urlopen(request, timeout=15):
                sent = True
        except Exception:
            pass

    bot_token = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    chat_id = os.getenv("TELEGRAM_CHAT_ID", "").strip()
    if bot_token and chat_id:
        try:
            tg_url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
            payload = json.dumps({"chat_id": chat_id, "text": message, "parse_mode": "Markdown"}).encode("utf-8")
            req = urllib.request.Request(tg_url, data=payload, headers={"Content-Type": "application/json"}, method="POST")
            with urllib.request.urlopen(req, timeout=15):
                sent = True
        except Exception:
            pass

    return sent


def _append_trade_event(track: dict, event_type: str, price: float, now: str) -> None:
    """Store one lifecycle event per ticker/event type, and optionally push it once."""
    seen = set(track.setdefault("eventTypes", []))
    if event_type in seen:
        return
    entry = float(track.get("startPrice") or price)
    high = float(track.get("highestPrice") or price)
    target_index = {"TP1": 0, "TP2": 1, "TP3": 2}.get(event_type)
    target = (track.get("targets") or [None, None, None])[target_index] if target_index is not None else track.get("effectiveStop")
    realized = round((price / entry - 1) * 100, 2) if entry else 0.0
    maximum = round((high / entry - 1) * 100, 2) if entry else 0.0
    actions = {
        "TP1": "TP1 görüldü: stop maliyete çekildi; yeni hedefler izleniyor.",
        "TP2": "TP2 görüldü: stop TP1 seviyesine çekildi; yeni hedef izleniyor.",
        "TP3": "TP3 görüldü: pozisyon planı tamamlandı; kademeli kâr alma değerlendirilebilir.",
        "TRAILING_STOP": "Kârlı stop tetiklendi: korunan kârla takip sona erdi.",
        "STOP": "Stop seviyesi görüldü: plan dışına çıkmadan işlem kapatıldı.",
    }
    title = {"TP1": "TP1 HEDEFİ GÖRÜLDÜ", "TP2": "TP2 HEDEFİ GÖRÜLDÜ", "TP3": "TP3 HEDEFİ GÖRÜLDÜ", "TRAILING_STOP": "KÂRLI STOP TETİKLENDİ", "STOP": "STOP TETİKLENDİ"}[event_type]
    message = f"{title}\n{track.get('ticker')} · {track.get('strategy') or 'Model'}\nGiriş: {entry:.2f} TL · Gerçekleşen: {price:.2f} TL · Getiri: {realized:+.2f}%\n{actions[event_type]}"
    delivered = _notify(message)
    row = {
        "timestamp": now, "ticker": track.get("ticker"), "price": price,
        "strategy": track.get("strategy"), "score": track.get("score"),
        "message": message, "status": "sent" if delivered else "local",
        "eventType": event_type, "eventTitle": title, "entryPrice": entry,
        "targetPrice": target, "signalAt": track.get("startedAt"),
        "realizedAt": now, "elapsedSeconds": max(0, int((datetime.fromisoformat(now) - datetime.fromisoformat(track["startedAt"])).total_seconds())),
        "realizedReturn": realized, "maxPrice": high, "maxReturn": maximum,
        "action": actions[event_type],
    }
    try:
        rows = json.loads(NOTIFICATION_LOG_PATH.read_text(encoding="utf-8")) if NOTIFICATION_LOG_PATH.exists() else []
        rows = rows if isinstance(rows, list) else []
        rows.insert(0, row)
        NOTIFICATION_LOG_PATH.write_text(json.dumps(rows[:500], ensure_ascii=False), encoding="utf-8")
    except (OSError, ValueError, TypeError):
        pass
    track["eventTypes"].append(event_type)
    track.setdefault("events", []).append(row)


def _advance_trade_lifecycle(track: dict, price: float, now: str) -> None:
    """Advance TP/stop state from delayed prices. It never creates or executes an order."""
    if track.get("closed"):
        return
    targets = [float(value) for value in (track.get("targets") or []) if value not in (None, "")]
    entry = float(track.get("startPrice") or price)
    stop = float(track.get("initialStop") or track.get("effectiveStop") or 0)
    for index, target in enumerate(targets[:3]):
        event = f"TP{index + 1}"
        if price >= target and event not in set(track.get("eventTypes", [])):
            _append_trade_event(track, event, price, now)
            if index == 0:
                track["effectiveStop"] = max(stop, entry)
            elif index == 1:
                track["effectiveStop"] = max(float(track.get("effectiveStop") or stop), targets[0])
            else:
                track["effectiveStop"] = max(float(track.get("effectiveStop") or stop), targets[1] if len(targets) > 1 else entry)
    effective_stop = float(track.get("effectiveStop") or stop)
    if effective_stop > 0 and price <= effective_stop:
        protected = any(event in set(track.get("eventTypes", [])) for event in ("TP1", "TP2", "TP3"))
        _append_trade_event(track, "TRAILING_STOP" if protected and effective_stop >= entry else "STOP", price, now)
        track["closed"] = True


def opportunity_snapshot(payload: dict) -> list[dict]:
    opportunities = []
    for stock in payload.get("stocks", []):
        delayed = stock.get("delayedQuote") or {}
        if stock.get("recommendation") != "OPEN" or not delayed.get("price"):
            continue
        low, high = stock.get("entryZoneLow"), stock.get("entryZoneHigh")
        price = float(delayed["price"])
        if low is None or high is None or not (float(low) <= price <= float(high)):
            continue
        notes = stock.get("analystNotes", [])
        analyst_message = next((note.get("Alarm Açıklaması / Talimatı") or note.get("Özel Açıklamalar / Analiz Notları") for note in notes if note.get("Alarm Açıklaması / Talimatı") or note.get("Özel Açıklamalar / Analiz Notları")), None)
        opportunities.append({"ticker": stock["ticker"], "price": price, "score": stock.get("modelScore"), "strategy": stock.get("strategy"), "entry": [low, high], "stop": stock.get("stop"), "targets": stock.get("targets"), "analystNotes": notes, "analystMessage": analyst_message})
    return opportunities


def manual_opportunity_snapshot(payload: dict) -> list[dict]:
    selected = set(manual_tracking())
    rows = []
    for stock in payload.get("stocks", []):
        if stock.get("ticker") not in selected:
            continue
        quote = stock.get("delayedQuote") or {}
        price = quote.get("price") or stock.get("price")
        if not price:
            continue
        notes = stock.get("analystNotes", [])
        analyst_message = next((note.get("Alarm Açıklaması / Talimatı") or note.get("Özel Açıklamalar / Analiz Notları") for note in notes if note.get("Alarm Açıklaması / Talimatı") or note.get("Özel Açıklamalar / Analiz Notları")), None)
        rows.append({"ticker": stock["ticker"], "price": float(price), "score": stock.get("modelScore"), "strategy": "Manuel analist takibi", "entry": [stock.get("entryZoneLow") or price, stock.get("entryZoneHigh") or price], "stop": stock.get("stop"), "targets": stock.get("targets") or [price], "analystMessage": analyst_message, "manual": True})
    return rows


def _load_analyst_levels() -> list[dict]:
    """Load analyst support/resistance/entry levels from Excel and OGUZ_ANALIZ_ARSIVI.json."""
    levels = []

    # 1. Load from OGUZ_ANALIZ_ARSIVI.json
    try:
        if OGUZ_ARSIV_PATH.exists():
            rows = json.loads(OGUZ_ARSIV_PATH.read_text(encoding="utf-8"))
            for row in rows:
                ticker = row.get("ticker", "").upper().strip()
                if not ticker:
                    continue
                entry = row.get("entry_level")
                note = row.get("note", "")
                source = "Oğuz Çelik" if "Oğuz" in note else "Analist"
                if entry:
                    levels.append({"ticker": ticker, "source": source, "entry": float(entry), "support": None, "resistance": None, "note": note})
    except Exception:
        pass

    # 2. Load from Excel
    try:
        if EXCEL_PATH.exists():
            import openpyxl
            wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
            ws = wb.active
            for r in range(4, ws.max_row + 1):
                ticker = ws.cell(row=r, column=2).value
                if not ticker:
                    continue
                ticker = str(ticker).upper().strip()
                name = ws.cell(row=r, column=3).value or ""
                entry = ws.cell(row=r, column=4).value
                support_raw = ws.cell(row=r, column=5).value
                resistance_raw = ws.cell(row=r, column=6).value
                note = ws.cell(row=r, column=8).value or ""

                source = "Oğuz Çelik" if "Oğuz" in note else "Ahmet Mergen" if "Mergen" in note else "Analist"

                # Parse support/resistance (may be strings like "1.34 - 1.35 TL")
                support_val = None
                resistance_val = None
                try:
                    if support_raw:
                        s = str(support_raw).replace("TL", "").replace(",", ".").strip()
                        parts = [p.strip() for p in s.split("-") if p.strip()]
                        support_val = float(parts[0]) if parts else None
                except Exception:
                    pass
                try:
                    if resistance_raw:
                        s = str(resistance_raw).replace("TL", "").replace(",", ".").strip()
                        parts = [p.strip() for p in s.split("-") if p.strip()]
                        resistance_val = float(parts[-1]) if parts else None
                except Exception:
                    pass

                # Only add if we don't already have this ticker from OGUZ or if Excel has extra data
                existing_tickers = {l["ticker"] for l in levels}
                if ticker not in existing_tickers:
                    levels.append({"ticker": ticker, "source": source, "entry": float(entry) if entry else None, "support": support_val, "resistance": resistance_val, "note": note, "name": name})
                else:
                    # Merge Excel support/resistance into existing
                    for l in levels:
                        if l["ticker"] == ticker:
                            if support_val and not l.get("support"):
                                l["support"] = support_val
                            if resistance_val and not l.get("resistance"):
                                l["resistance"] = resistance_val
                            if note and "Mergen" in note and "Mergen" not in (l.get("note") or ""):
                                l["note"] = l.get("note", "") + " | " + note
                            break
            wb.close()
    except Exception:
        pass

    return levels


def check_analyst_level_alerts(payload: dict) -> None:
    """Check if any stock price is near analyst support/resistance/entry levels and send detailed alerts."""
    analyst_levels = _load_analyst_levels()
    if not analyst_levels:
        return

    state = _load_state()
    today_str = get_tr_now().strftime("%Y-%m-%d")
    changed = False

    stocks_map = {}
    for s in payload.get("stocks", []):
        quote = s.get("delayedQuote") or {}
        price = quote.get("price") or s.get("price")
        if price:
            stocks_map[s["ticker"]] = {"price": float(price), "stock": s}

    for level in analyst_levels:
        ticker = level["ticker"]
        if ticker not in stocks_map:
            continue

        price = stocks_map[ticker]["price"]
        stock = stocks_map[ticker]["stock"]
        entry = level.get("entry")
        support = level.get("support")
        resistance = level.get("resistance")
        note = level.get("note") or ""
        source = level.get("source", "Analist")
        name = level.get("name") or stock.get("name") or ticker

        # Proximity threshold: within 2% of a key level
        threshold = 0.02
        alerts = []

        if entry and abs(price - entry) / entry <= threshold:
            direction = "📗 Giriş seviyesinin TAM ÜZERİNDE" if price >= entry else "📕 Giriş seviyesinin ALTINA düştü"
            diff_pct = round((price / entry - 1) * 100, 2)
            alerts.append({
                "type": "GİRİŞ SEVİYESİ",
                "level": entry,
                "direction": direction,
                "diff_pct": diff_pct,
                "emoji": "🎯"
            })

        if support and price <= support * (1 + threshold):
            diff_pct = round((price / support - 1) * 100, 2)
            if diff_pct <= 2:
                direction = "🟢 Destek seviyesinde TUTUNUYOR" if price >= support else "🔴 Destek seviyesi KIRILDI!"
                alerts.append({
                    "type": "DESTEK SEVİYESİ",
                    "level": support,
                    "direction": direction,
                    "diff_pct": diff_pct,
                    "emoji": "🛡️"
                })

        if resistance and price >= resistance * (1 - threshold):
            diff_pct = round((price / resistance - 1) * 100, 2)
            if diff_pct >= -2:
                direction = "🟢 Direnç seviyesi KIRILDI! Yukarı yön açıldı" if price >= resistance else "🟡 Direnç seviyesine YAKIN, satıcı gelebilir"
                alerts.append({
                    "type": "DİRENÇ SEVİYESİ",
                    "level": resistance,
                    "direction": direction,
                    "diff_pct": diff_pct,
                    "emoji": "🧱"
                })

        for alert in alerts:
            alert_key = f"analyst_{ticker}_{alert['type']}_{today_str}"
            if state.get(alert_key):
                continue

            # Build a super detailed message
            model_score = stock.get("modelScore", "—")
            targets = stock.get("targets") or [price, price, price]
            tp1 = targets[0] if len(targets) > 0 else price
            tp2 = targets[1] if len(targets) > 1 else tp1
            tp3 = targets[2] if len(targets) > 2 else tp2
            stop = stock.get("stop", "—")

            message = (
                f"📢 *ANALİST SEVİYE ALARMI*\n"
                f"👤 *Kaynak:* {source}\n\n"
                f"📌 *Hisse:* #{ticker} ({name})\n"
                f"💵 *Güncel Fiyat:* {price:.2f} TL\n"
                f"⭐ *Model Puanı:* {model_score}\n\n"
                f"{alert['emoji']} *{alert['type']}:* {alert['level']:.2f} TL\n"
                f"{alert['direction']}\n"
                f"📊 *Fark:* %{alert['diff_pct']:+.2f}\n\n"
                f"💬 *Analist Notu:*\n{note}\n\n"
                f"🎯 *Model Hedefleri:*\n"
                f"  • TP1: {tp1} TL\n"
                f"  • TP2: {tp2} TL\n"
                f"  • TP3: {tp3} TL\n"
                f"  • Stop: {stop} TL\n\n"
                f"⚠️ *Ne Yapmalı:*\n"
            )

            if alert["type"] == "DESTEK SEVİYESİ":
                if price >= support:
                    message += f"Destek {alert['level']:.2f} TL'de tutuyor. Buradan alım düşünülebilir ama stop {alert['level']:.2f} TL altına konmalı. Kırılırsa uzak dur."
                else:
                    message += f"Destek {alert['level']:.2f} TL kırıldı! Pozisyon varsa stop'u değerlendir. Yeni alım için acele etme, daha aşağı gelebilir."
            elif alert["type"] == "DİRENÇ SEVİYESİ":
                if price >= resistance:
                    message += f"Direnç {alert['level']:.2f} TL kırıldı! Yukarı yön açıldı. Kâr hedeflerine doğru izle. Geri dönerse dikkat."
                else:
                    message += f"Direnç {alert['level']:.2f} TL'ye yaklaştı. Burada satıcı gelebilir. Pozisyon varsa kısmi kâr alma düşünülebilir."
            else:
                message += f"Giriş seviyesi {alert['level']:.2f} TL civarında. Analist bu seviyeyi alım için uygun görmüştü. Stop koyarak değerlendirilebilir."

            delivered = _notify(message)
            if delivered:
                state[alert_key] = True
                changed = True

    if changed:
        _save_state(state)


def check_and_send_scheduled_summaries(payload: dict) -> None:
    now_dt = get_tr_now()
    today_str = now_dt.strftime("%Y-%m-%d")
    hour = now_dt.hour
    minute = now_dt.minute

    # Determine current slot for TR Local Hours (UTC+3): 10:00, 12:00, 14:00, 16:00, 18:00, 18:30
    slot_name = None
    slot_key = None
    if hour == 10 and minute <= 30:
        slot_name = "10:00 (Seans Açılış Özet Bülteni)"
        slot_key = f"summary_{today_str}_1000"
    elif hour == 12 and minute <= 30:
        slot_name = "12:00 (Seans Ortası Özet Bülteni)"
        slot_key = f"summary_{today_str}_1200"
    elif hour == 14 and minute <= 30:
        slot_name = "14:00 (Öğleden Sonra Özet Bülteni)"
        slot_key = f"summary_{today_str}_1400"
    elif hour == 16 and minute <= 30:
        slot_name = "16:00 (Kapanış Öncesi Özet Bülteni)"
        slot_key = f"summary_{today_str}_1600"
    elif hour == 18 and minute <= 15:
        slot_name = "18:00 (Seans Kapanış ve Gün Sonu Bülteni)"
        slot_key = f"summary_{today_str}_1800"

    if not slot_name:
        return

    state = _load_state()
    if state.get(slot_key):
        return  # Already sent for this specific slot today

    index_info = payload.get("index", {})
    xu100_price = index_info.get("price", "—")
    xu100_daily = index_info.get("daily", 0.0)

    index_warning = ""
    try:
        p_val = float(xu100_price)
        if p_val <= 13850:
            index_warning = "\n⚠️ *DİKKAT:* Endeks 13.850 TL kritik desteğinde! Aşağı düşerse sonraki durak 12.900 – 13.000 TL olabilir. Yükseliş için 14.250 TL üzerine çıkmalı."
        else:
            index_warning = f"\nℹ️ *Endeks Durumu:* {p_val:,.2f} TL (13.850 TL kritik destektir; altına düşerse 12.900 TL riski doğar, 14.250 TL üzeri alım teyididir)."
    except Exception:
        index_warning = ""

    wind_text = f"BIST100: {xu100_price} TL (%{xu100_daily:+.2f}){index_warning}"

    stocks = payload.get("stocks", [])
    top_candidates = sorted(stocks, key=lambda s: s.get("modelScore", 0), reverse=True)[:5]
    
    top_text_list = []
    for s in top_candidates:
        targets = s.get("targets") or [s["price"], s["price"], s["price"]]
        tp1 = targets[0] if len(targets) > 0 else s["price"]
        tp2 = targets[1] if len(targets) > 1 else tp1
        tp3 = targets[2] if len(targets) > 2 else tp2
        stop = s.get("stop", "—")
        badge = s.get("badges", [""])[0] if s.get("badges") else ""
        badge_str = f" ({badge})" if badge else ""
        
        top_text_list.append(
            f"📌 *#{s['ticker']}*{badge_str} — Model Puanı: *{s['modelScore']}*\n"
            f"  • Güncel Fiyat: {s['price']} TL | Stop: {stop} TL\n"
            f"  • Kâr Hedefleri: TP1: {tp1} TL | TP2: {tp2} TL | TP3: {tp3} TL"
        )

    top_text = "\n\n".join(top_text_list) if top_text_list else "Fırsat hisse bulunamadı."

    try:
        from auto_portfolio import load_portfolio
        portfolio = load_portfolio()
        initial = portfolio.get("initial_capital", 10000.0)
        cash = portfolio.get("current_cash", 10000.0)
        positions = portfolio.get("positions", [])
        equity_val = sum(pos.get("current_price", 0) * pos.get("qty", 0) for pos in positions)
        total_val = cash + equity_val
        net_pnl = total_val - initial
        net_pnl_pct = (net_pnl / initial) * 100

        portfolio_text = (
            f"• Toplam Bakiye: {total_val:,.2f} TL\n"
            f"• Boştaki Nakit: {cash:,.2f} TL ({len(positions)} Aktif Pozisyon)\n"
            f"• Net K/Z: {net_pnl:+.2f} TL (%{net_pnl_pct:+.2f})"
        )
    except Exception:
        portfolio_text = "Portföy bilgisi alınamadı."

    message = (
        f"🕒 *MODEL EN YÜKSEK PUAN ÖZET BÜLTENİ*\n"
        f"📅 *Zaman:* {today_str} — {slot_name}\n\n"
        f"📈 *Piyasa Yönü:*\n{wind_text}\n\n"
        f"🤖 *10K Robot Portföy Karnesi:*\n{portfolio_text}\n\n"
        f"⭐ *Modeldeki En Yüksek Puanlı Top 5 Hisse:*\n\n{top_text}\n\n"
        f"🏷️ *Sözlük:* SK3 = 3'lü Süper Konsensüs | ÇAO = Çifte Algo Onayı"
    )

    if _notify(message):
        state[slot_key] = True
        _save_state(state)


def check_intraday_price_movements(payload: dict) -> None:
    """Check if any tracked stock has significant intraday price movements (e.g. >= 2% gain/loss) and push alerts."""
    state = _load_state()
    today_str = get_tr_now().strftime("%Y-%m-%d")
    changed = False

    for stock in payload.get("stocks", []):
        ticker = stock.get("ticker")
        quote = stock.get("delayedQuote") or {}
        price = quote.get("price") or stock.get("price")
        prev_close = (stock.get("officialOhlc") or {}).get("previousClose") or stock.get("previousClose") or price

        if not ticker or not price or not prev_close or float(prev_close) <= 0:
            continue

        price = float(price)
        prev_close = float(prev_close)
        change_pct = round((price / prev_close - 1) * 100, 2)

        # Triggers for +2.0%, +4.0%, +6.0%, and -3.0%
        levels = []
        if change_pct >= 6.0:
            levels.append(("+6%", "🚀 *GÜÇLÜ RALLİ HAREKETİ*"))
        elif change_pct >= 4.0:
            levels.append(("+4%", "🔥 *SEANS İÇİ İVME KAZANDI*"))
        elif change_pct >= 2.0:
            levels.append(("+2%", "📈 *POZİTİF YÜKSELİŞ HAREKETİ*"))
        elif change_pct <= -3.0:
            levels.append(("-3%", "⚠️ *SEANS İÇİ GERİ ÇEKİLME*"))

        for pct_tag, title in levels:
            alert_key = f"move_{ticker}_{pct_tag}_{today_str}"
            if state.get(alert_key):
                continue

            model_score = stock.get("modelScore", "—")
            notes = stock.get("analystNotes", [])
            note_str = next((n.get("Alarm Açıklaması / Talimatı") or n.get("Özel Açıklamalar / Analiz Notları") for n in notes if n.get("Alarm Açıklaması / Talimatı") or n.get("Özel Açıklamalar / Analiz Notları")), "Model Takibi")

            message = (
                f"{title}\n\n"
                f"📌 *Hisse:* #{ticker}\n"
                f"💵 *Güncel Fiyat:* {price:.2f} TL (Günlük: %{change_pct:+.2f})\n"
                f"⭐ *Model Puanı:* {model_score}\n"
                f"💬 *Analist / Strateji Notu:*\n{note_str}\n\n"
                f"📊 *Önceki Kapanış:* {prev_close:.2f} TL"
            )

            delivered = _notify(message)
            if delivered:
                state[alert_key] = True
                changed = True

    if changed:
        _save_state(state)


def run_once() -> list[dict]:
    try:
        payload = scan_market()
        opportunities = opportunity_snapshot(payload)
        existing = {item["ticker"] for item in opportunities}
        opportunities.extend(item for item in manual_opportunity_snapshot(payload) if item["ticker"] not in existing)
        tracks = _record_tracking(payload, opportunities)
        state = _load_state()
        today_str = get_tr_now().strftime("%Y-%m-%d")
        changed = False
        for item in opportunities:
            key = f"{today_str}|manual|{item['ticker']}" if item.get("manual") else f"{today_str}|{item['ticker']}|{item['strategy']}"
            if state.get(key):
                continue

            stock_obj = next((s for s in payload.get("stocks", []) if s.get("ticker") == item["ticker"]), None)
            is_manual = item.get("manual", False)
            score = item.get("score", 0) or 0

            consensus_badges = []
            if stock_obj and stock_obj.get("badges"):
                # Filter positive consensus badges (SK3, ÇAO), excluding risk tags
                consensus_badges = [b for b in stock_obj["badges"] if ("SK" in b or "ÇAO" in b) and "RİSK" not in b and "OBO" not in b]

            badge_info = ""
            if consensus_badges:
                badge_info = f"\n🏷️ *KONSENSÜS ROZETİ:* {consensus_badges[0]}"

            is_super = any("SK" in b for b in consensus_badges)
            is_double = any("ÇAO" in b for b in consensus_badges)

            if is_manual:
                header = f"📢 *ANALİST TAKİP ALARMI*"
            elif is_super:
                header = f"🚀 *SÜPER KONSENSÜS ALARMI*"
            elif is_double:
                header = f"🔥 *ÇİFTE ALGO ONAY ALARMI*"
            else:
                header = f"⚡ *MODEL FIRSAT ALARMI*"

            entry_low = float(item['entry'][0])
            entry_high = float(item['entry'][1])
            raw_stop = float(item.get('stop') or (entry_low * 0.96))
            effective_stop = round(min(raw_stop, entry_low * 0.97), 2)
            if effective_stop >= entry_low:
                effective_stop = round(entry_low * 0.96, 2)

            targets = item.get("targets") or [item["price"], item["price"], item["price"]]
            tp1 = targets[0] if len(targets) > 0 else item["price"]
            tp2 = targets[1] if len(targets) > 1 else tp1
            tp3 = targets[2] if len(targets) > 2 else tp2

            message = (
                f"{header}\n\n"
                f"📌 *Hisse:* #{item['ticker']}\n"
                f"⭐ *Model Puanı:* {item['score']}\n"
                f"💬 *Açıklama:* {item.get('analystMessage') or item['strategy']}"
                f"{badge_info}\n\n"
                f"💵 *Güncel Fiyat:* {item['price']} TL\n"
                f"🎯 *Alım Bölgesi:* {entry_low}–{entry_high} TL arası\n"
                f"🛑 *Zarar Kes (Stop):* {effective_stop} TL (Altına düşerse satıp çıkılmalı)\n"
                f"🚀 *Kâr Hedefleri:*\n"
                f"  • *1. Hedef (TP1):* {tp1} TL\n"
                f"  • *2. Hedef (TP2):* {tp2} TL\n"
                f"  • *3. Hedef (TP3):* {tp3} TL"
            )
            delivered = _notify(message)
            _append_notification_log(item, message, delivered)
            if delivered:
                state[key] = True
                changed = True
        if changed:
            _save_state(state)

        # Always check and send scheduled summary bulletins for 10:00, 12:00, 14:00, 16:00, 18:00, 18:30
        check_and_send_scheduled_summaries(payload)

        # Check Oğuz/Mergen analyst support/resistance/entry level proximity alerts
        check_analyst_level_alerts(payload)

        # Check real-time intraday price movements (e.g. +2%, +4%, +6% moves)
        check_intraday_price_movements(payload)

        # Update 10K Paper Trading Auto-Portfolio state automatically
        try:
            from auto_portfolio import update_auto_portfolio
            update_auto_portfolio(payload.get("stocks", []), today_str)
        except Exception:
            pass

        with _status_lock:
            _status.update({"running": True, "lastRun": datetime.now().astimezone().isoformat(timespec="seconds"), "lastError": None, "opportunities": opportunities, "tracking": list(tracks.values()), "notificationsEnabled": bool(os.getenv("NTFY_TOPIC", DEFAULT_NTFY_TOPIC).strip()), "topic": os.getenv("NTFY_TOPIC", DEFAULT_NTFY_TOPIC).strip()})
        return opportunities
    except Exception as exc:
        with _status_lock:
            _status.update({"running": True, "lastRun": datetime.now().astimezone().isoformat(timespec="seconds"), "lastError": str(exc)})
        raise


def worker_status() -> dict:
    with _status_lock:
        return dict(_status)


def notification_log(limit: int = 200) -> list[dict]:
    try:
        rows = json.loads(NOTIFICATION_LOG_PATH.read_text(encoding="utf-8")) if NOTIFICATION_LOG_PATH.exists() else []
        if not isinstance(rows, list):
            return []
        deduped = []
        failed_tickers = set()
        for row in rows:
            if row.get("status") == "not_sent":
                ticker = row.get("ticker")
                if ticker in failed_tickers:
                    continue
                failed_tickers.add(ticker)
            deduped.append(row)
        return deduped[: max(1, min(int(limit), 500))]
    except (OSError, ValueError, TypeError):
        return []


def tracking_log(limit: int = 200) -> list[dict]:
    rows = list(_load_tracking().values())
    rows.sort(key=lambda item: item.get("lastAt", ""), reverse=True)
    return rows[: max(1, min(int(limit), 500))]


def _load_tracking() -> dict[str, dict]:
    try:
        value = json.loads(TRACKING_LOG_PATH.read_text(encoding="utf-8")) if TRACKING_LOG_PATH.exists() else {}
        return value if isinstance(value, dict) else {}
    except (OSError, ValueError, TypeError):
        return {}


def _save_tracking(value: dict[str, dict]) -> None:
    TRACKING_LOG_PATH.write_text(json.dumps(value, ensure_ascii=False), encoding="utf-8")


def _record_tracking(payload: dict, opportunities: list[dict]) -> dict[str, dict]:
    tracks = _load_tracking()
    now = datetime.now().astimezone().isoformat(timespec="seconds")
    new_by_ticker = {item["ticker"]: item for item in opportunities}
    stock_by_ticker = {item.get("ticker"): item for item in payload.get("stocks", [])}
    for ticker, item in new_by_ticker.items():
        quote = stock_by_ticker.get(ticker, {}).get("delayedQuote") or {}
        price = float(quote.get("price") or item.get("price"))
        if ticker not in tracks:
            tracks[ticker] = {
                "ticker": ticker,
                "strategy": item.get("strategy"),
                "score": item.get("score"),
                "analystMessage": item.get("analystMessage"),
                "startedAt": now,
                "startPrice": price,
                "lastAt": now,
                "lastPrice": price,
                "changePercent": 0.0,
                "targets": item.get("targets") or [],
                "initialStop": item.get("stop"),
                "effectiveStop": item.get("stop"),
                "highestPrice": price,
                "highestAt": now,
                "eventTypes": [],
                "events": [],
                "closed": False,
                "updates": [{"timestamp": now, "price": price}],
            }
        elif item.get("analystMessage") and not tracks[ticker].get("analystMessage"):
            tracks[ticker]["analystMessage"] = item.get("analystMessage")
    for ticker, track in tracks.items():
        stock = stock_by_ticker.get(ticker)
        if not stock:
            continue
        quote = stock.get("delayedQuote") or {}
        price = quote.get("price") or stock.get("price")
        try:
            price = float(price)
        except (TypeError, ValueError):
            continue
        if price <= 0:
            continue
        if price != float(track.get("lastPrice", 0)):
            track.setdefault("updates", []).append({"timestamp": now, "price": price})
            track["updates"] = track["updates"][-100:]
        track["lastAt"] = now
        track["lastPrice"] = price
        if price >= float(track.get("highestPrice") or price):
            track["highestPrice"] = price
            track["highestAt"] = now
        start = float(track.get("startPrice") or price)
        track["changePercent"] = round((price / start - 1) * 100, 2) if start else 0.0
        _advance_trade_lifecycle(track, price, now)
    _save_tracking(tracks)
    return tracks


if __name__ == "__main__":
    while True:
        try:
            print(json.dumps(run_once(), ensure_ascii=False))
        except Exception as exc:
            print(json.dumps({"error": str(exc)}, ensure_ascii=False))
        time.sleep(INTERVAL_SECONDS)
