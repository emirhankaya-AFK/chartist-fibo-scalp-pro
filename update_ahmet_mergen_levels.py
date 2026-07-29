import openpyxl
import json

# Ahmet Mergen Live Stream Analysis Data (2026)
mergen_stocks = [
    {
        "ticker": "USAK",
        "name": "Uşak Seramik Sanayi A.Ş.",
        "entry": 1.37,
        "live_price": 1.37,
        "support": "1.34 - 1.35 TL",
        "resistance": "1.80 TL",
        "note": "[Ahmet Mergen] 2024 ve 2025 eski dipleri 1.35-1.37 TL seviyesine geldi. Dolarda 2 cent dip seviyesi. Ortalamaların üzerine çıkması ve hacim beklenmeli."
    },
    {
        "ticker": "YKBNK",
        "name": "Yapı ve Kredi Bankası A.Ş.",
        "entry": 32.00,
        "live_price": 32.00,
        "support": "28.00 - 30.00 TL (Dolarda 60 Cent)",
        "resistance": "36.00 TL",
        "note": "[Ahmet Mergen] Bankalarda 50 aylık ortalama test ediliyor. Dolarda 60 cent seviyesi kritik ana destek. 60 cent kırılırsa XBANK 2026 riski doğar."
    },
    {
        "ticker": "MOGAN",
        "name": "Mogan Enerji Yatırım A.Ş.",
        "entry": 19.68,
        "live_price": 17.63,
        "support": "13.75 TL (50 Günlük HO)",
        "resistance": "22.52 TL",
        "note": "[Ahmet Mergen] Eski alım yeri 7.50 TL'ydi. 22.52 TL büyük çanak tamamlanma direnci. Fiyat ortalamadan açıldığı için kâr satışı normal."
    },
    {
        "ticker": "PAHOL",
        "name": "Pasifik Holding A.Ş.",
        "entry": 1.40,
        "live_price": 1.40,
        "support": "1.25 - 1.32 TL",
        "resistance": "1.67 TL",
        "note": "[Ahmet Mergen] Yatay alandan aşağı sarkıyor. 1.32-1.35 ve 1.25 TL desteklerine süzülüyor. 50 günlük ortalama altında alım yönü zayıf."
    },
    {
        "ticker": "SAHOL",
        "name": "Sabancı Holding A.Ş.",
        "entry": 85.00,
        "live_price": 85.00,
        "support": "75.00 - 80.00 TL",
        "resistance": "105.00 - 110.00 TL",
        "note": "[Ahmet Mergen] 75-105 TL devasa yatay bant. 85 TL 4. dip bölgesi, esas ana toplama yeri 75 TL. 105 TL kırılırsa katlama hedefi 130-135 TL."
    },
    {
        "ticker": "NUGYO",
        "name": "Nurol GMYO A.Ş.",
        "entry": 8.70,
        "live_price": 8.70,
        "support": "7.50 - 8.50 TL",
        "resistance": "12.00 - 14.00 TL",
        "note": "[Ahmet Mergen] Oynaklığı yüksek spekülatif tahta. 8.50 TL civarı alıcı bölgeleri. 12 TL üzerinde satıcı gelir."
    },
    {
        "ticker": "NATEN",
        "name": "Natürel Yenilenebilir Enerji A.Ş.",
        "entry": 6.00,
        "live_price": 6.00,
        "support": "5.00 - 5.50 TL",
        "resistance": "8.50 TL",
        "note": "[Ahmet Mergen] 1.5 yıldır 50 haftalık ortalama altında düşüşte. Mumlar küçülüyor, 5.00-5.50 TL dip bölgesine yanaşma sinyali var."
    },
    {
        "ticker": "TUPRS",
        "name": "Tüpraş A.Ş.",
        "entry": 292.25,
        "live_price": 292.25,
        "support": "275.00 - 280.00 TL",
        "resistance": "320.00 - 335.00 TL",
        "note": "[Ahmet Mergen] 4 haftada %50 yükseldi, tepede Doji oluştu. Petrol geri çekilmesiyle 275-280 TL 20 günlük HO desteğine süzülebilir. Çanak katlama hedefi 330-335 TL."
    },
    {
        "ticker": "PETKM",
        "name": "Petkim Petrokimya A.Ş.",
        "entry": 20.00,
        "live_price": 20.00,
        "support": "18.50 TL",
        "resistance": "23.50 TL",
        "note": "[Ahmet Mergen] Petrol sert yükselişiyle %50 primlendi, brent petrol düzeltmesiyle ortalamalara çekilme riski mevcut."
    }
]

# 1. Update Hisselerin_Teknik_Verileri.xlsx
wb = openpyxl.load_workbook('Hisselerin_Teknik_Verileri.xlsx')
ws = wb.active

existing_rows = {}
for r in range(4, ws.max_row + 1):
    val = ws.cell(row=r, column=2).value
    if val:
        ticker_str = str(val).upper().strip()
        existing_rows[ticker_str] = r

for stock in mergen_stocks:
    ticker = stock["ticker"]
    if ticker in existing_rows:
        row_idx = existing_rows[ticker]
        ws.cell(row=row_idx, column=4).value = stock["live_price"]
        ws.cell(row=row_idx, column=5).value = stock["support"]
        ws.cell(row=row_idx, column=6).value = stock["resistance"]
        ws.cell(row=row_idx, column=8).value = stock["note"]
    else:
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = ticker
        ws.cell(row=new_row, column=2).value = ticker
        ws.cell(row=new_row, column=3).value = stock["name"]
        ws.cell(row=new_row, column=4).value = stock["live_price"]
        ws.cell(row=new_row, column=5).value = stock["support"]
        ws.cell(row=new_row, column=6).value = stock["resistance"]
        ws.cell(row=new_row, column=8).value = stock["note"]

wb.save('Hisselerin_Teknik_Verileri.xlsx')
print("Hisselerin_Teknik_Verileri.xlsx updated with Mergen analysis.")

# 2. Update manual_tracking.json
with open("manual_tracking.json", "r", encoding="utf-8") as f:
    manual_list = json.load(f)

for stock in mergen_stocks:
    if stock["ticker"] not in manual_list:
        manual_list.append(stock["ticker"])

with open("manual_tracking.json", "w", encoding="utf-8") as f:
    json.dump(manual_list, f, ensure_ascii=False, indent=2)
print("manual_tracking.json updated.")
