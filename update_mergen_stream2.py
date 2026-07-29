import openpyxl
import json

new_mergen_stocks = [
    {
        "ticker": "RYSAS",
        "name": "Reysaş Taşımacılık ve Lojistik A.Ş.",
        "entry": 24.00,
        "support": "21.64 TL (50 Günlük HO)",
        "resistance": "25.00 TL - 26.50 TL",
        "note": "[Ahmet Mergen] 20 TL altından 24 TL'ye tepki verdi. 25.00 ve 26.50 TL ara dirençler. Çanak tamamlama potansiyeli mevcut. 50 günlük HO (21.64 TL) iz süren stop."
    },
    {
        "ticker": "TTRAK",
        "name": "Türk Traktör ve Ziraat Makineleri A.Ş.",
        "entry": 400.00,
        "support": "400.00 TL (Dolar: 5.5-6 USD)",
        "resistance": "500.00 TL (50 Günlük HO)",
        "note": "[Ahmet Mergen] 1.000 TL'den 400 TL'ye süzüldü. %80'e yakın çekilme yaşandı. 500 TL (50 günlük HO) seviyesine %20'lik tepki hareketi yapabilir."
    },
    {
        "ticker": "HALKB",
        "name": "Türkiye Halk Bankası A.Ş.",
        "entry": 29.50,
        "support": "29.00 - 30.00 TL (0.72 - 0.75 USD)",
        "resistance": "34.00 TL",
        "note": "[Ahmet Mergen] Dolar bazında 50 aylık ortalama (0.72-0.75$) dip bölgesine geldi. BIST bankacılık genel baskısı altında."
    },
    {
        "ticker": "VAKBN",
        "name": "Türkiye Vakıflar Bankası T.A.O.",
        "entry": 19.50,
        "support": "18.50 - 19.00 TL",
        "resistance": "22.00 TL",
        "note": "[Ahmet Mergen] 50 aylık ortalama dip bölgesini test ediyor. Bankacılık dip kırılımları takibinde."
    },
    {
        "ticker": "PGSUS",
        "name": "Pegasus Hava Taşımacılığı A.Ş.",
        "entry": 154.00,
        "support": "135.00 - 150.00 TL",
        "resistance": "200.00 TL - 240.00 TL",
        "note": "[Ahmet Mergen] Petrol sıçraması ve jeopolitik riskle 285 TL'den 150 TL'ye (%50) sert geriledi. 135-150 TL dip bölgesi."
    },
    {
        "ticker": "DOAS",
        "name": "Doğuş Otomotiv Servis ve Ticaret A.Ş.",
        "entry": 195.00,
        "support": "175.00 - 180.00 TL",
        "resistance": "210.00 TL",
        "note": "[Ahmet Mergen] Kısa vadede aşağı kayıyor. 175-180 TL yatay ana dip desteği. 195 TL üzerine çıkamazsa 180 TL'ye süzülebilir."
    },
    {
        "ticker": "SISE",
        "name": "Türkiye Şişe ve Cam Fabrikaları A.Ş.",
        "entry": 42.00,
        "support": "40.00 - 42.00 TL",
        "resistance": "48.00 TL",
        "note": "[Ahmet Mergen] 42 TL seviyelerinde dip arayışında, baskı devam ediyor."
    }
]

# 1. Update Hisselerin_Teknik_Verileri.xlsx
wb = openpyxl.load_workbook('Hisselerin_Teknik_Verileri.xlsx')
ws = wb.active

existing_rows = {}
for r in range(4, ws.max_row + 1):
    val = ws.cell(row=r, column=2).value
    if val:
        ticker_str = str(val).upper().strip()
        existing_rows[ticker_str] = r

for stock in new_mergen_stocks:
    ticker = stock["ticker"]
    if ticker in existing_rows:
        row_idx = existing_rows[ticker]
        ws.cell(row=row_idx, column=4).value = stock["entry"]
        ws.cell(row=row_idx, column=5).value = stock["support"]
        ws.cell(row=row_idx, column=6).value = stock["resistance"]
        ws.cell(row=row_idx, column=8).value = stock["note"]
    else:
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = ticker
        ws.cell(row=new_row, column=2).value = ticker
        ws.cell(row=new_row, column=3).value = stock["name"]
        ws.cell(row=new_row, column=4).value = stock["entry"]
        ws.cell(row=new_row, column=5).value = stock["support"]
        ws.cell(row=new_row, column=6).value = stock["resistance"]
        ws.cell(row=new_row, column=8).value = stock["note"]

wb.save('Hisselerin_Teknik_Verileri.xlsx')
print("Hisselerin_Teknik_Verileri.xlsx updated with second Mergen stream.")

# 2. Update manual_tracking.json
with open("manual_tracking.json", "r", encoding="utf-8") as f:
    manual_list = json.load(f)

for stock in new_mergen_stocks:
    if stock["ticker"] not in manual_list:
        manual_list.append(stock["ticker"])

with open("manual_tracking.json", "w", encoding="utf-8") as f:
    json.dump(manual_list, f, ensure_ascii=False, indent=2)
print("manual_tracking.json updated.")
