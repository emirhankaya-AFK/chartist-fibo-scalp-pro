import openpyxl
import json

# Verified live prices and extracted video metadata
oguz_stocks = [
    {
        "ticker": "AKENR",
        "name": "Ak Enerji Elektrik Üretim A.Ş.",
        "entry": 11.00,
        "live_price": 10.56,
        "note": "[Oğuz Analizi] Akenr 11₺ değerlendirip satanlar için. 10.38₺ takibi."
    },
    {
        "ticker": "KRONT",
        "name": "Kron Telekomünikasyon Hizmetleri",
        "entry": 22.16,
        "live_price": 24.26,
        "note": "[Oğuz Analizi] 22.16₺ ve 22.80₺ bakılabilir, %5 direnç bölgesi."
    },
    {
        "ticker": "YAYLA",
        "name": "Yayla Enerji Üretim Turizm",
        "entry": 23.10,
        "live_price": 22.98,
        "note": "[Oğuz Analizi] 23.10₺ ve 23.78₺ tradelik bakılabilir, 24₺ hareketi."
    },
    {
        "ticker": "DOFRB",
        "name": "Dof Robotik Sanayi A.Ş.",
        "entry": 75.00,
        "live_price": 151.80,
        "note": "[Oğuz Analizi] 75₺'den 90₺ ve 107.60₺ seviyesine hareket."
    },
    {
        "ticker": "MEYSU",
        "name": "Meysu Gıda Sanayi A.Ş.",
        "entry": 13.03,
        "live_price": 13.77,
        "note": "[Oğuz Analizi] Meysu 13.03₺ tradelik bakıyorum (%6.5 marj)."
    },
    {
        "ticker": "EKOS",
        "name": "Ekos Teknoloji ve Elektrik A.Ş.",
        "entry": 6.00,
        "live_price": 6.03,
        "note": "[Oğuz Analizi] Ekos 6₺ seviyesinden bakılabilir."
    },
    {
        "ticker": "LILAK",
        "name": "Lila Kağıt Sanayi ve Ticaret",
        "entry": 30.80,
        "live_price": 30.16,
        "note": "[Oğuz Analizi] Lilak 30.80₺'de iyi duruyor (%2 güzel marj)."
    },
    {
        "ticker": "OBASE",
        "name": "Obase Bilgisayar ve Danışmanlık",
        "entry": 47.98,
        "live_price": 44.44,
        "note": "[Oğuz Analizi] Obase 47.98₺ seviyesinden takibe alındı."
    },
    {
        "ticker": "NETCD",
        "name": "Netcad Yazılım A.Ş.",
        "entry": 144.50,
        "live_price": 139.30,
        "note": "[Oğuz Analizi] Netcd 144.50₺ takibe alındı (%3.5 marj)."
    },
    {
        "ticker": "ASTOR",
        "name": "Astor Enerji A.Ş.",
        "entry": 286.00,
        "live_price": 297.00,
        "note": "[Oğuz Analizi] 286₺ Astor bakılabilir, 295₺ ve 299₺ dirençleri."
    },
    {
        "ticker": "IZMDC",
        "name": "İzmir Demir Çelik Sanayi A.Ş.",
        "entry": 6.156,
        "live_price": 12.16,
        "note": "[Oğuz Analizi] Izmdc 6.156₺ buradan takip ediyorum (%110 prim)."
    },
    {
        "ticker": "AKHAN",
        "name": "Akhan Un Fabrikası A.Ş.",
        "entry": 23.80,
        "live_price": 40.16,
        "note": "[Oğuz Analizi] 23.80₺ / 25.12₺ seviyesinden tarım emtiası takibi."
    },
    {
        "ticker": "ZERGY",
        "name": "Zergay Gayrimenkul Yatırım A.Ş.",
        "entry": 10.06,
        "live_price": 10.24,
        "note": "[Oğuz Analizi] Zergy 10.06₺ takip, 10.50₺ direnç bölgesi."
    },
    {
        "ticker": "MARMR",
        "name": "Marmara Holding A.Ş.",
        "entry": 2.22,
        "live_price": 2.23,
        "note": "[Oğuz Analizi] Marmr 2.22₺ yine bakılır (%8 marj)."
    },
    {
        "ticker": "GWIND",
        "name": "Galata Wind Enerji A.Ş.",
        "entry": 24.36,
        "live_price": 25.38,
        "note": "[Oğuz Analizi] Gwind 24.36₺ benzer marj beklenebilir (%7.5)."
    },
    {
        "ticker": "KLRHO",
        "name": "Kiler Holding A.Ş.",
        "entry": 83.00,
        "live_price": 81.15,
        "note": "[Oğuz Analizi] Klrho 83₺ yine tradelik bakılabilir (%3.5)."
    },
    {
        "ticker": "KFEIN",
        "name": "Kafein Yazılım Hizmetleri A.Ş.",
        "entry": 9.40,
        "live_price": 8.92,
        "note": "[Oğuz Analizi] Kfein 9.40₺ seviyesinden takibe alındı."
    },
    {
        "ticker": "SSAAT",
        "name": "Saat ve Saat Sanayi A.Ş.",
        "entry": 42.24,
        "live_price": 46.16,
        "note": "[Oğuz Analizi] SSAAT 42.24₺ takip seviyesi."
    },
    {
        "ticker": "RUZYE",
        "name": "Ruzy Madencilik A.Ş.",
        "entry": 9.20,
        "live_price": 9.50,
        "note": "[Oğuz Analizi] Ruzye 9.20₺ anlık ağır yükseldi, tradelik bakılır."
    },
    {
        "ticker": "OZATD",
        "name": "Ozfatura Teknoloji A.Ş.",
        "entry": 151.00,
        "live_price": 3120.00,
        "note": "[Oğuz Analizi] 151 TL'den alınıp orta-uzun vade takip edilen analiz."
    },
    {
        "ticker": "TEHOL",
        "name": "Tek-Art Holding A.Ş.",
        "entry": 43.24,
        "live_price": 43.10,
        "note": "[Oğuz Analizi] %140.84 net kâr ile kapatılan analiz."
    },
    {
        "ticker": "GENKM",
        "name": "Gentaş Kimya Sanayi A.Ş.",
        "entry": 12.25,
        "live_price": 12.18,
        "note": "[Oğuz Analizi] Genkm 12.25₺ takip, yetişemeyene %4 marj."
    },
    {
        "ticker": "UCAYM",
        "name": "Ucay Mühendislik A.Ş.",
        "entry": 28.80,
        "live_price": 28.46,
        "note": "[Oğuz Analizi] 29.20₺ üzeri takip konusu."
    },
    {
        "ticker": "BULGS",
        "name": "Bulls Girişim Sermayesi A.Ş.",
        "entry": 39.48,
        "live_price": 39.10,
        "note": "[Oğuz Analizi] Bulgs 39.48₺ tradelik takip edilebilir (%3.5)."
    },
    {
        "ticker": "EKDMR",
        "name": "Ekinciler Demir Çelik A.Ş.",
        "entry": 54.10,
        "live_price": 51.65,
        "note": "[Oğuz Analizi] Ekinciler Demir Çelik 54.10₺ takibe alındı."
    },
    {
        "ticker": "MOGAN",
        "name": "Mogan Enerji Yatırım A.Ş.",
        "entry": 12.75,
        "live_price": 17.63,
        "note": "[Oğuz Analizi] Mogan 12.75₺ buradan pozisyon açıyor (%60 prim)."
    },
    {
        "ticker": "KCAER",
        "name": "Kocaer Çelik Sanayi A.Ş.",
        "entry": 11.14,
        "live_price": 15.00,
        "note": "[Oğuz Analizi] Kcaer 11.14₺ tradelik değil değerlenme (%45)."
    },
    {
        "ticker": "DSTKF",
        "name": "Destek Faktoring A.Ş.",
        "entry": 2180.00,
        "live_price": 2048.00,
        "note": "[Oğuz Analizi] Destek Faktoring tradelik takip (Giriş ~2.180 TL, Anlık ~2.048 TL)."
    }
]

# 1. Update Hisselerin_Teknik_Verileri.xlsx
wb = openpyxl.load_workbook('Hisselerin_Teknik_Verileri.xlsx')
ws = wb.active

# Find existing tickers in Excel
existing_rows = {}
for r in range(4, ws.max_row + 1):
    val = ws.cell(row=r, column=2).value
    if val:
        ticker_str = str(val).upper().strip()
        existing_rows[ticker_str] = r

for stock in oguz_stocks:
    ticker = stock["ticker"]
    if ticker in existing_rows:
        row_idx = existing_rows[ticker]
        ws.cell(row=row_idx, column=4).value = stock["live_price"]
        ws.cell(row=row_idx, column=8).value = stock["note"]
    else:
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1).value = ticker
        ws.cell(row=new_row, column=2).value = ticker
        ws.cell(row=new_row, column=3).value = stock["name"]
        ws.cell(row=new_row, column=4).value = stock["live_price"]
        ws.cell(row=new_row, column=8).value = stock["note"]

wb.save('Hisselerin_Teknik_Verileri.xlsx')
print("Hisselerin_Teknik_Verileri.xlsx updated successfully.")

# 2. Update OGUZ_ANALIZ_ARSIVI.json
archive_data = []
for stock in oguz_stocks:
    ret_pct = round(((stock["live_price"] / stock["entry"]) - 1) * 100, 2)
    archive_data.append({
        "ticker": stock["ticker"],
        "name": stock["name"],
        "entry_level": stock["entry"],
        "current_live_price": stock["live_price"],
        "performance": f"{ret_pct:+.2f}%",
        "note": stock["note"]
    })

with open("OGUZ_ANALIZ_ARSIVI.json", "w", encoding="utf-8") as f:
    json.dump(archive_data, f, ensure_ascii=False, indent=2)
print("OGUZ_ANALIZ_ARSIVI.json updated successfully.")

# 3. Update manual_tracking.json
manual_list = [s["ticker"] for s in oguz_stocks]
with open("manual_tracking.json", "w", encoding="utf-8") as f:
    json.dump(manual_list, f, ensure_ascii=False, indent=2)
print("manual_tracking.json updated successfully.")
